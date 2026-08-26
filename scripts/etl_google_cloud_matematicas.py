#!/usr/bin/env python3
"""Carga idempotente del Excel de Google Cloud a academia-matematicas-soa.

Por seguridad el modo predeterminado es auditoria (dry-run). Use --apply para
escribir. Antes de cualquier escritura se genera un respaldo JSON comprimido.
"""

from __future__ import annotations

import argparse
import gzip
import html
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import pymysql
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Faltan dependencias. Ejecute: python -m pip install pymysql openpyxl"
    ) from exc


TARGET_DATABASE = "academia-matematicas-soa"
MANAGED_MARKER = "etl-google-cloud-matematicas"
SECTION_COLUMNS = {
    5: ("html", "Leccion escrita"),
    6: ("example", "Ejemplo resuelto"),
    7: ("activity", "Actividad"),
    8: ("evaluation", "Evaluacion"),
    9: ("activity", "Juego con plataforma"),
    10: ("video", "Video-leccion"),
    11: ("presentation", "Presentacion"),
    12: ("activity", "Actividad interactiva"),
}
COLORS = ["purple", "blue", "green", "orange", "pink", "teal", "yellow"]


def parse_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    required = ["DB_HOST", "DB_PORT", "DB_NAME", "DB_USER", "DB_PASSWORD"]
    missing = [key for key in required if not values.get(key)]
    if missing:
        raise ValueError(f"Faltan variables en .env: {', '.join(missing)}")
    if values["DB_NAME"] != TARGET_DATABASE:
        raise ValueError(
            f"Destino rechazado: se esperaba {TARGET_DATABASE!r}, "
            f"pero .env contiene {values['DB_NAME']!r}"
        )
    return values


def connect(env: dict[str, str], *, autocommit: bool = False):
    return pymysql.connect(
        host=env["DB_HOST"],
        port=int(env["DB_PORT"]),
        user=env["DB_USER"],
        password=env["DB_PASSWORD"],
        database=env["DB_NAME"],
        charset="utf8mb4",
        connect_timeout=10,
        read_timeout=120,
        write_timeout=120,
        autocommit=autocommit,
        cursorclass=pymysql.cursors.DictCursor,
    )


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def slugify(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", clean(value))
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", ascii_text)).strip("-")


def key_slug(value: Any) -> str:
    slug = slugify(value)
    if not re.fullmatch(r"[a-z]{3}-(?:p[1-6]|s[1-3]|pre[1-3])-\d{3}", slug):
        raise ValueError(f"Clave unica invalida: {value!r}")
    return slug


def education_level_for_grade(grade: str) -> str:
    if grade.startswith("PRE"):
        return "preparatoria"
    if grade.startswith("S"):
        return "secundaria"
    return "primaria"


def canonical_branch(value: Any) -> str:
    branch = clean(value)
    aliases = {
        "Estadistica": "Estadística",
        "Prealgebra?": "Preálgebra",
    }
    return aliases.get(branch, branch)


def category_slug(branch: str, education_level: str = "primaria") -> str:
    # Aritmetica ya es una categoria oficial del sitio y debe reutilizarse.
    if slugify(branch) == "aritmetica" and education_level == "primaria":
        return "aritmetica"
    prefix = "curriculo-" if education_level == "primaria" else f"curriculo-{education_level}-"
    return prefix + slugify(branch)


def as_html(value: Any, marker: str) -> str:
    text = str(value).strip()
    prefix = f"<!-- {MANAGED_MARKER}:{marker} -->\n"
    if re.search(r"<(?:!doctype|html|section|div|p|h[1-6])\b", text, re.I):
        return prefix + text
    return prefix + "<p>" + html.escape(text).replace("\n", "<br>\n") + "</p>"


def load_excel(path: Path) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    required = {"Lista de temas", "Lecciones creadas", "Multimedias"}
    missing = sorted(required - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"Faltan hojas requeridas: {', '.join(missing)}")

    topics: list[dict[str, Any]] = []
    topic_keys: set[str] = set()
    topic_by_number: dict[int, str] = {}
    for row_number, row in enumerate(
        workbook["Lista de temas"].iter_rows(min_row=2, values_only=True), start=2
    ):
        if row[0] in (None, ""):
            continue
        code = key_slug(row[7] if len(row) > 7 and row[7] else f"{str(row[2])[:3]}-{row[1]}-{int(row[0]):03d}")
        if code in topic_keys:
            raise ValueError(f"Clave duplicada en Lista de temas, fila {row_number}: {code}")
        topic_keys.add(code)
        topic_by_number[int(row[0])] = code
        grade = clean(row[1]).upper()
        if not re.fullmatch(r"(?:P[1-6]|S[1-3]|PRE[1-3])", grade):
            raise ValueError(f"Grado invalido en fila {row_number}: {grade!r}")
        topics.append(
            {
                "code": code,
                "source_number": int(row[0]),
                "grade": grade,
                "education_level": education_level_for_grade(grade),
                "branch": canonical_branch(row[2]),
                "area": clean(row[3]),
                "title": clean(row[4]) or clean(row[3]) or code,
                "unit": clean(row[5]),
                "subtopics": clean(row[6]),
            }
        )

    created: dict[str, list[dict[str, Any]]] = defaultdict(list)
    orphan_created: list[str] = []
    for row_number, row in enumerate(
        workbook["Lecciones creadas"].iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row[0]:
            continue
        try:
            code = key_slug(row[0])
        except ValueError:
            orphan_created.append(f"fila {row_number}: {row[0]!r}")
            continue
        if code not in topic_keys:
            orphan_created.append(f"fila {row_number}: {code}")
            continue
        for index, (section_type, label) in SECTION_COLUMNS.items():
            if index >= len(row) or row[index] in (None, ""):
                continue
            marker = f"created:{row_number}:{index}"
            created[code].append(
                {
                    "type": section_type,
                    "title": label,
                    "body": as_html(row[index], marker),
                    # Las columnas de "Lecciones creadas" ya contienen material
                    # terminado. Deben quedar visibles para el estudiante.
                    "published": 1,
                    "source": marker,
                }
            )

    media_plans: list[dict[str, Any]] = []
    orphan_media: list[str] = []
    for row_number, row in enumerate(
        workbook["Multimedias"].iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row[0]:
            continue
        try:
            code = key_slug(row[0])
        except ValueError:
            orphan_media.append(f"fila {row_number}: {row[0]!r}")
            continue
        if code not in topic_keys:
            # En el libro hay formulas de clave desactualizadas en algunas filas
            # multimedia. El numero final es la clave estable de Lista de temas.
            number_match = re.search(r"(\d{1,3})$", code)
            source_number = int(number_match.group(1)) if number_match else -1
            corrected = topic_by_number.get(source_number)
            if not corrected:
                orphan_media.append(f"fila {row_number}: {code}")
                continue
            code = corrected
        media_type = clean(row[3]).lower()
        section_type = "video" if "video" in media_type else "presentation" if "present" in media_type else "activity"
        marker = f"media-plan:{row_number}"
        media_plans.append(
            {
                "code": code,
                "type": section_type,
                "title": clean(row[1]) or clean(row[3]) or "Recurso multimedia",
                "body": as_html(row[2] or row[3], marker),
                "published": 0,
                "source": marker,
            }
        )

    audit = {
        "workbook_sheets": [
            {"name": ws.title, "rows": ws.max_row, "columns": ws.max_column}
            for ws in workbook.worksheets
        ],
        "topic_count": len(topics),
        "created_section_count": sum(map(len, created.values())),
        "media_plan_count": len(media_plans),
        "grades": dict(Counter(t["grade"] for t in topics)),
        "branches": dict(Counter(t["branch"] for t in topics)),
        "orphan_created": orphan_created,
        "orphan_media": orphan_media,
    }
    return topics, created, media_plans, audit


def inspect_sql_dump(path: Path) -> dict[str, Any]:
    """Audita el dump sin ejecutar DDL destructivo ni bases ajenas."""
    databases: list[str] = []
    target_tables: list[str] = []
    target_inserts: Counter[str] = Counter()
    active = False
    with path.open("r", encoding="utf-8", errors="replace") as source:
        for line in source:
            use = re.match(r"USE `([^`]+)`", line)
            if use:
                database = use.group(1)
                databases.append(database)
                active = database == TARGET_DATABASE
                continue
            if not active:
                continue
            create = re.match(r"CREATE TABLE `([^`]+)`", line)
            if create:
                target_tables.append(create.group(1))
            insert = re.match(r"INSERT INTO `([^`]+)`", line)
            if insert:
                target_inserts[insert.group(1)] += 1
    return {
        "databases_found": list(dict.fromkeys(databases)),
        "target_tables": target_tables,
        "target_insert_statements": dict(target_inserts),
        "note": "El dump se audita, pero no se ejecuta porque contiene DROP TABLE y otras cinco bases ajenas.",
    }


def json_default(value: Any):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (bytes, bytearray)):
        return {"__bytes_hex__": bytes(value).hex()}
    raise TypeError(type(value).__name__)


def backup_database(connection, destination: Path) -> Path:
    destination.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output = destination / f"{TARGET_DATABASE}-antes-etl-{stamp}.json.gz"
    payload: dict[str, Any] = {
        "database": TARGET_DATABASE,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "tables": {},
    }
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema=%s ORDER BY table_name",
            (TARGET_DATABASE,),
        )
        names = [row["TABLE_NAME"] for row in cursor.fetchall()]
        for table in names:
            cursor.execute(f"SHOW CREATE TABLE `{table}`")
            create_row = cursor.fetchone()
            cursor.execute(f"SELECT * FROM `{table}`")
            payload["tables"][table] = {
                "create_sql": create_row["Create Table"],
                "rows": cursor.fetchall(),
            }
    with gzip.open(output, "wt", encoding="utf-8") as target:
        json.dump(payload, target, ensure_ascii=False, default=json_default)
    if output.stat().st_size == 0:
        raise RuntimeError("El respaldo quedo vacio")
    return output


def current_counts(connection) -> dict[str, int]:
    result: dict[str, int] = {}
    with connection.cursor() as cursor:
        for table in [
            "aprendizaje_categorias",
            "aprendizaje_lecciones",
            "aprendizaje_secciones_leccion",
            "aprendizaje_medios",
        ]:
            cursor.execute(f"SELECT COUNT(*) AS total FROM `{table}`")
            result[table] = int(cursor.fetchone()["total"])
    return result


def apply_etl(connection, topics, created, media_plans) -> dict[str, int]:
    branches = list(dict.fromkeys((t["education_level"], t["branch"]) for t in topics))
    stats: Counter[str] = Counter()
    lesson_ids: dict[str, int] = {}
    try:
        with connection.cursor() as cursor:
            for order, (education_level, branch) in enumerate(branches, start=10):
                branch_slug = category_slug(branch, education_level)
                cursor.execute(
                    """INSERT INTO aprendizaje_categorias
                    (name,slug,description,education_level,icon,color,sort_order,active)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,1)
                    ON DUPLICATE KEY UPDATE name=VALUES(name),description=VALUES(description),
                    education_level=VALUES(education_level),color=VALUES(color),
                    sort_order=VALUES(sort_order),active=VALUES(active)""",
                    (
                        branch,
                        branch_slug,
                        "Rama curricular importada desde la estructura de matematicas de Google Cloud.",
                        education_level,
                        "#",
                        COLORS[(order - 10) % len(COLORS)],
                        order,
                    ),
                )
                stats["categories_upserted"] += 1

            cursor.execute(
                "SELECT id,slug FROM aprendizaje_categorias "
                "WHERE slug LIKE 'curriculo-%' OR slug='aritmetica'"
            )
            categories = {row["slug"]: row["id"] for row in cursor.fetchall()}

            for topic in topics:
                category_id = categories[category_slug(topic["branch"], topic["education_level"])]
                summary_parts = [
                    f"Clave curricular: {topic['code']}",
                    f"Grado: {topic['grade']}",
                    f"Rama: {topic['branch']}",
                ]
                if topic["area"]:
                    summary_parts.append(f"Area: {topic['area']}")
                if topic["unit"]:
                    summary_parts.append(f"Unidad: {topic['unit']}")
                if topic["subtopics"]:
                    summary_parts.append(f"Subtemas: {topic['subtopics']}")
                summary = ". ".join(summary_parts) + "."
                cursor.execute(
                    """INSERT INTO aprendizaje_lecciones
                    (category_id,title,slug,summary,page_type,icon,icon_type,
                     difficulty,duration_minutes,sort_order,published)
                    VALUES (%s,%s,%s,%s,'topic',%s,'emoji','Basica',30,%s,1)
                    ON DUPLICATE KEY UPDATE category_id=VALUES(category_id),title=VALUES(title),
                    summary=VALUES(summary),sort_order=VALUES(sort_order),published=VALUES(published)""",
                    (
                        category_id,
                        f"{topic['grade']} - {topic['title']}",
                        topic["code"],
                        summary,
                        "#",
                        topic["source_number"],
                    ),
                )
                cursor.execute(
                    "SELECT id FROM aprendizaje_lecciones WHERE slug=%s", (topic["code"],)
                )
                lesson_ids[topic["code"]] = int(cursor.fetchone()["id"])
                stats["lessons_upserted"] += 1

            managed_ids = list(lesson_ids.values())
            if managed_ids:
                placeholders = ",".join(["%s"] * len(managed_ids))
                cursor.execute(
                    f"DELETE FROM aprendizaje_secciones_leccion "
                    f"WHERE lesson_id IN ({placeholders}) AND body_html LIKE %s",
                    (*managed_ids, f"<!-- {MANAGED_MARKER}:%"),
                )
                stats["old_managed_sections_deleted"] = cursor.rowcount

            sections_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for code, sections in created.items():
                sections_by_code[code].extend(sections)
            for item in media_plans:
                sections_by_code[item["code"]].append(item)

            for code, sections in sections_by_code.items():
                for order, section in enumerate(sections, start=1):
                    cursor.execute(
                        """INSERT INTO aprendizaje_secciones_leccion
                        (lesson_id,section_type,title,body_html,ai_exercises_enabled,
                         duration_minutes,sort_order,published)
                        VALUES (%s,%s,%s,%s,0,0,%s,%s)""",
                        (
                            lesson_ids[code],
                            section["type"],
                            section["title"][:180],
                            section["body"],
                            order,
                            section["published"],
                        ),
                    )
                    stats["sections_inserted"] += 1

            # Limpieza especifica de una categoria transitoria creada por la
            # primera version del ETL. Solo se borra si ya no tiene lecciones.
            cursor.execute(
                "DELETE c FROM aprendizaje_categorias c "
                "LEFT JOIN aprendizaje_lecciones l ON l.category_id=c.id "
                "WHERE c.slug='curriculo-aritmetica' AND l.id IS NULL"
            )
            stats["empty_transitional_categories_deleted"] = cursor.rowcount
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    return dict(stats)


def main() -> int:
    project = Path(__file__).resolve().parents[1]
    workspace = project.parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="Escribir en MySQL")
    parser.add_argument("--env", type=Path, default=project / ".env")
    parser.add_argument(
        "--source-dir", type=Path, default=workspace / "Bases de Datos Google Cloud"
    )
    parser.add_argument("--report", type=Path, default=project / "database" / "etl-reports")
    parser.add_argument("--backup-dir", type=Path, default=project / "database" / "backups")
    args = parser.parse_args()

    excel_files = sorted(args.source_dir.glob("*.xlsx"))
    dump = args.source_dir / "respaldo_completo.sql"
    if len(excel_files) != 1:
        raise ValueError(f"Se esperaba exactamente un XLSX; encontrados: {len(excel_files)}")
    if not dump.is_file():
        raise FileNotFoundError(dump)

    env = parse_env(args.env)
    topics, created, media_plans, excel_audit = load_excel(excel_files[0])
    report: dict[str, Any] = {
        "mode": "apply" if args.apply else "dry-run",
        "target_database": TARGET_DATABASE,
        "source_excel": str(excel_files[0]),
        "source_dump": str(dump),
        "excel": excel_audit,
        "sql_dump": inspect_sql_dump(dump),
    }
    connection = connect(env)
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DATABASE() AS db, VERSION() AS version")
            report["server"] = cursor.fetchone()
        report["counts_before"] = current_counts(connection)
        if args.apply:
            backup = backup_database(connection, args.backup_dir)
            report["backup"] = str(backup)
            report["changes"] = apply_etl(connection, topics, created, media_plans)
            report["counts_after"] = current_counts(connection)
        else:
            report["planned"] = {
                "categories_upsert": len(set((t["education_level"], t["branch"]) for t in topics)),
                "lessons_upsert": len(topics),
                "sections_replace": sum(map(len, created.values())) + len(media_plans),
                "lesson_html_sections_published": sum(map(len, created.values())),
                "multimedia_plans_published": False,
            }
    finally:
        connection.close()

    args.report.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output = args.report / f"etl-google-cloud-{report['mode']}-{stamp}.json"
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2, default=json_default))
    print(f"\nReporte: {output}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
